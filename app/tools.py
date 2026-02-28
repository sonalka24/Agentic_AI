import json
import re
import base64
from pathlib import Path
from datetime import datetime, timezone
from urllib import parse, request

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from langchain_openai import ChatOpenAI
except Exception:
    ChatOpenAI = None

from config import Config


def load_prompts(prompts_path=None):
    """Load prompt templates from JSON and merge with defaults."""
    path = prompts_path or Path(__file__).with_name("prompts.json")
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data
    except Exception:
        raise ValueError('Could not load prompts.json')

    

class Toolset:
    """Collection of data + LLM tools used by the LangGraph agent."""

    def __init__(self, config, minio_client):
        """Initialize toolset dependencies and prompt registry.

        @param config Runtime configuration object.
        @param minio_client Initialized MinIO client instance.
        @return None.
        """
        self.config = config
        self.client = minio_client
        self._llm = None
        self.prompts = load_prompts()

    def _get_llm(self):
        """Create/reuse a ChatOpenAI client; return None when unavailable."""
        if self._llm is not None:
            return self._llm
        if not self.config.openai_api_key or ChatOpenAI is None:
            return None
        try:
            self._llm = ChatOpenAI(
                api_key=self.config.openai_api_key,
                model=self.config.openai_model,
                temperature=0,
            )
            return self._llm
        except Exception:
            return None

    @staticmethod
    def _normalize_cell_value(value):
        """Normalize a spreadsheet cell value to string.

        @param value Raw cell value from pandas/openpyxl.
        @return str|None Trimmed string or None for empty/NaN values.
        """
        if pd.isna(value):
            return None
        if hasattr(value, "isoformat") and not isinstance(value, str):
            try:
                return value.isoformat()
            except Exception:
                pass
        return str(value).strip()

    @staticmethod
    def _safe_json_load(text):
        """Parse JSON with relaxed substring fallback.

        @param text JSON-like text.
        @return dict|list|None Parsed object or None on failure.
        """
        try:
            return json.loads(text)
        except Exception:
            start = text.find("{")
            end = text.rfind("}")
            if start != -1 and end != -1 and end > start:
                try:
                    return json.loads(text[start : end + 1])
                except Exception:
                    return None
            return None

    @staticmethod
    def _rows_to_text(rows):
        """Convert row arrays into numbered pipe-delimited text lines.

        @param rows list[list[str]] Row values.
        @return str Serialized text block.
        """
        lines = []
        for i, row in enumerate(rows):
            lines.append(f"{i + 1}. " + " | ".join([str(x) for x in row]))
        return "\n".join(lines)

    @staticmethod
    def _text_to_rows(sheet_text):
        """Parse numbered pipe-delimited sheet text back into rows.

        @param sheet_text Serialized sheet text.
        @return list[list[str]] Parsed row values.
        """
        rows = []
        for line in str(sheet_text).splitlines():
            raw = line.strip()
            if not raw:
                continue
            if ". " in raw:
                prefix, rest = raw.split(". ", 1)
                if prefix.isdigit():
                    raw = rest
            cells = [c.strip() for c in raw.split("|")]
            cells = [c for c in cells if c]
            if cells:
                rows.append(cells)
        return rows

    @staticmethod
    def _contains_chinese(text):
        """Detect whether text contains CJK ideographs.

        @param text Input text.
        @return bool True when Chinese characters are present.
        """
        return bool(re.search(r"[\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF]", str(text)))

    @staticmethod
    def _extract_chinese_tokens(text):
        """Extract unique contiguous Chinese token sequences.

        @param text Input text.
        @return list[str] Unique tokens in first-seen order.
        """
        tokens = re.findall(r"[\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF]+", str(text))
        unique_tokens = []
        seen = set()
        for token in tokens:
            if token not in seen:
                seen.add(token)
                unique_tokens.append(token)
        return unique_tokens

    @staticmethod
    def _replace_tokens(text, token_map):
        """Replace tokens in text using longest-token-first order.

        @param text Source text.
        @param token_map Mapping from original token to replacement token.
        @return str Replaced text.
        """
        out = str(text)
        for token in sorted(token_map.keys(), key=len, reverse=True):
            replacement = str(token_map.get(token, ""))
            if replacement:
                out = out.replace(token, replacement)
        return out

    def _clickhouse_http_query(self, query, body=None):
        """Execute an HTTP SQL request against ClickHouse.

        @param query SQL query string.
        @param body Optional request payload for INSERT statements.
        @return str Raw response body.
        @raises RuntimeError On transport/query failures.
        """
        host = self.config.clickhouse_host
        port = self.config.clickhouse_port
        db = self.config.clickhouse_db
        params = parse.urlencode({"database": db, "query": query})
        url = f"http://{host}:{port}/?{params}"

        req = request.Request(
            url=url,
            data=body.encode("utf-8") if isinstance(body, str) else body,
            method="POST",
        )
        user = str(self.config.clickhouse_user or "")
        password = str(self.config.clickhouse_password or "")
        token = base64.b64encode(f"{user}:{password}".encode("utf-8")).decode("utf-8")
        req.add_header("Authorization", f"Basic {token}")
        req.add_header("Content-Type", "text/plain; charset=utf-8")
        try:
            with request.urlopen(req, timeout=30) as resp:
                return resp.read().decode("utf-8", errors="ignore")
        except Exception as exc:
            raise RuntimeError(f"ClickHouse query failed: {exc}") from exc

    def _call_openai_json(self, prompt, raise_on_error=False):
        """Execute an LLM call expecting a JSON object and parse the result."""
        llm = self._get_llm()
        if llm is None:
            if raise_on_error:
                raise RuntimeError("OpenAI client unavailable. Check OPENAI_API_KEY/model configuration.")
            return None
        try:
            response = llm.bind(response_format={"type": "json_object"}).invoke(
                [
                    ("system", self.prompts["system_json_only"]),
                    ("user", prompt),
                ]
            )
            content = response.content if hasattr(response, "content") else ""
            if isinstance(content, list):
                content = "".join(
                    [str(x.get("text", x)) if isinstance(x, dict) else str(x) for x in content]
                )
            return self._safe_json_load(str(content))
        except Exception as exc:
            if raise_on_error:
                raise RuntimeError(f"OpenAI JSON call failed: {exc}") from exc
            return None

    def _translate_chinese_to_english(self, text):
        """Translate Chinese text to English while preserving row layout."""
        source = str(text)
        if not source.strip() or not self._contains_chinese(source):
            return source

        tokens = self._extract_chinese_tokens(source)
        if not tokens:
            return source

        base = self.prompts["translate_chinese_to_english"]
        prompt = (
            base
            + "\n\nTranslate only these Chinese tokens and return a mapping.\n"
            + "Tokens:\n"
            + json.dumps(tokens, ensure_ascii=False)
            + '\n\nReturn STRICT JSON with shape: {"translations":{"<chinese_token>":"<english_translation>"}}'
        )
        parsed = self._call_openai_json(prompt, raise_on_error=True)
        translations = parsed.get("translations") if isinstance(parsed, dict) else None
        if not isinstance(translations, dict):
            retry_prompt = (
                prompt
                + '\n\nYour previous output was invalid. Return ONLY: {"translations":{"...":"..."}}'
            )
            parsed = self._call_openai_json(retry_prompt, raise_on_error=True)
            translations = parsed.get("translations") if isinstance(parsed, dict) else None
        if not isinstance(translations, dict):
            raise RuntimeError("LLM translation failed: invalid JSON translations map.")

        token_map = {}
        for token in tokens:
            value = translations.get(token)
            if isinstance(value, str) and value.strip():
                token_map[token] = value.strip()

        if not token_map:
            raise RuntimeError("LLM translation failed: no token translations returned.")
        return self._replace_tokens(source, token_map)

    def _translate_sections_to_english(self, sections):
        """Translate section key/value payload text to English.

        @param sections Nested section/subsection/key map.
        @return dict Translated nested map preserving structure.
        """
        translated_sections = {}
        for section_name, subsection_map in sections.items():
            out_section = {}
            if not isinstance(subsection_map, dict):
                translated_sections[section_name] = out_section
                continue
            for subsection_name, kv in subsection_map.items():
                out_kv = {}
                if not isinstance(kv, dict):
                    out_section[subsection_name] = out_kv
                    continue
                for key, value in kv.items():
                    key_text = str(key)
                    value_text = str(value) if value is not None else ""
                    key_en = (
                        self._translate_chinese_to_english(key_text)
                        if self._contains_chinese(key_text)
                        else key_text
                    )
                    value_en = (
                        self._translate_chinese_to_english(value_text)
                        if self._contains_chinese(value_text)
                        else value_text
                    )
                    out_kv[key_en] = value_en
                out_section[subsection_name] = out_kv
            translated_sections[section_name] = out_section
        return translated_sections

    @staticmethod
    def _normalize_sections_from_llm(parsed, target_section_schema):
        """Normalize LLM output to the expected section/subsection schema.

        @param parsed Raw parsed JSON from LLM response.
        @param target_section_schema Required section/subsection schema.
        @return tuple[dict|None, str|None] Normalized sections and product_id.
        """
        if not isinstance(parsed, dict):
            return None, None

        candidate = parsed
        if not isinstance(candidate.get("sections"), dict):
            for wrapper_key in ("data", "result", "output", "payload"):
                wrapped = candidate.get(wrapper_key)
                if isinstance(wrapped, dict) and isinstance(wrapped.get("sections"), dict):
                    candidate = wrapped
                    break

        raw_sections = candidate.get("sections")
        if not isinstance(raw_sections, dict):
            return None, None

        sections_filled = {}
        for section, subsection_schema in target_section_schema.items():
            raw_section = raw_sections.get(section)
            if raw_section is None:
                for sec_name, sec_payload in raw_sections.items():
                    if str(sec_name).strip().lower() == str(section).strip().lower():
                        raw_section = sec_payload
                        break
            if not isinstance(raw_section, dict):
                raw_section = {}

            # If model returned flat keys under section, place them into first subsection.
            has_nested = any(isinstance(v, dict) for v in raw_section.values())
            section_out = {}
            subsection_names = list(subsection_schema.keys())
            first_sub = subsection_names[0] if subsection_names else "General"

            if not has_nested:
                section_out[first_sub] = raw_section
                for sub in subsection_names:
                    section_out.setdefault(sub, {})
                sections_filled[section] = section_out
                continue

            for subsection in subsection_names:
                raw_kv = raw_section.get(subsection)
                if raw_kv is None:
                    for sub_name, sub_payload in raw_section.items():
                        if str(sub_name).strip().lower() == str(subsection).strip().lower():
                            raw_kv = sub_payload
                            break
                section_out[subsection] = raw_kv if isinstance(raw_kv, dict) else {}
            sections_filled[section] = section_out

        product_id = candidate.get("product_id", parsed.get("product_id", ""))
        return sections_filled, str(product_id) if product_id is not None else ""

    @staticmethod
    def _extract_product_id(rows, default_id):
        """Extract product identifier from parsed rows.

        @param rows Parsed row values.
        @param default_id Fallback identifier.
        @return str Product identifier.
        """
        for row in rows:
            if len(row) > 1 and str(row[0]).startswith("Art No."):
                return str(row[1])
        return default_id

    @staticmethod
    def _norm_label(text):
        """Normalize label text for tolerant matching.

        @param text Input label.
        @return str Lowercase alphanumeric-only key.
        """
        return re.sub(r"[^a-z0-9]+", "", str(text).strip().lower())

    @staticmethod
    def _safe_path_name(text):
        """Convert arbitrary text into filesystem-safe path segment.

        @param text Input label.
        @return str Safe path component.
        """
        cleaned = re.sub(r"[^\w.\-]+", "_", str(text).strip())
        return cleaned.strip("_") or "unknown"

    @staticmethod
    def _anchor_to_row_col(anchor):
        """Resolve openpyxl image anchor into 1-based row/column.

        @param anchor Image anchor object or A1 reference.
        @return tuple[int|None, int|None] Row and column coordinates.
        """
        # openpyxl image anchors are typically OneCellAnchor/TwoCellAnchor.
        marker = getattr(anchor, "_from", None)
        if marker is not None:
            return int(marker.row) + 1, int(marker.col) + 1
        if isinstance(anchor, str) and anchor:
            m = re.match(r"^([A-Za-z]+)(\d+)$", anchor)
            if m:
                col_letters, row_str = m.groups()
                col = 0
                for ch in col_letters.upper():
                    col = col * 26 + (ord(ch) - ord("A") + 1)
                return int(row_str), int(col)
        return None, None

    def _row_section_ranges(self, ws, target_sections):
        """Infer contiguous row ranges for each section header.

        @param ws Worksheet object.
        @param target_sections Ordered section names to detect.
        @return list[dict] Section ranges with start/end rows.
        """
        target_lookup = {}
        for section in target_sections:
            norm = self._norm_label(section)
            if norm:
                target_lookup[norm] = str(section)

        starts = []
        max_row = int(ws.max_row or 0)
        for row_idx, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
            row_tokens = []
            for value in row_values:
                normalized = self._normalize_cell_value(value)
                if normalized:
                    row_tokens.append(normalized)
            if not row_tokens:
                continue

            matched = None
            for token in row_tokens:
                token_norm = self._norm_label(token)
                if not token_norm:
                    continue
                if token_norm in target_lookup:
                    matched = target_lookup[token_norm]
                    break
                for section_norm, section_name in target_lookup.items():
                    if (
                        len(section_norm) >= 4
                        and (section_norm in token_norm or token_norm in section_norm)
                    ):
                        matched = section_name
                        break
                if matched:
                    break
            if matched:
                if not starts or starts[-1]["section"] != matched:
                    starts.append({"row": row_idx, "section": matched})

        if not starts:
            return []

        ranges = []
        for i, start in enumerate(starts):
            start_row = start["row"]
            end_row = starts[i + 1]["row"] - 1 if i + 1 < len(starts) else max_row
            ranges.append(
                {
                    "section": start["section"],
                    "start_row": start_row,
                    "end_row": end_row,
                }
            )
        return ranges

    @staticmethod
    def _resolve_section_for_row(row_num, section_ranges):
        """Resolve row number to nearest section range.

        @param row_num Image anchor row.
        @param section_ranges Section range metadata.
        @return str Matched section name or fallback label.
        """
        if row_num is None:
            return "Unassigned"
        for item in section_ranges:
            if item["start_row"] <= row_num <= item["end_row"]:
                return item["section"]
        # If image appears above first section, assign first; if below, assign last.
        if section_ranges and row_num < section_ranges[0]["start_row"]:
            return section_ranges[0]["section"]
        if section_ranges:
            return section_ranges[-1]["section"]
        return "Unassigned"

    def _assign_images_with_llm_and_fallback(self, images, section_rows, subsection_rows, target_section_schema):
        """Assign image anchors to section/subsection using LLM and deterministic fallback.

        @param images Image metadata rows.
        @param section_rows Detected section header rows.
        @param subsection_rows Detected subsection header rows.
        @param target_section_schema Allowed section/subsection schema.
        @return dict Mapping image_id -> assignment payload.
        """
        section_names = list(target_section_schema.keys())
        valid_pairs = set()
        for section, subsection_map in target_section_schema.items():
            if isinstance(subsection_map, dict):
                for subsection in subsection_map.keys():
                    valid_pairs.add((str(section), str(subsection)))

        section_rows_sorted = sorted(
            [r for r in section_rows if isinstance(r.get("row"), int)],
            key=lambda x: int(x["row"]),
        )
        subsection_rows_sorted = sorted(
            [r for r in subsection_rows if isinstance(r.get("row"), int)],
            key=lambda x: int(x["row"]),
        )

        def _section_from_row(image_row):
            if not section_rows_sorted:
                return "Unassigned"
            # Prefer the latest section header at or before image row.
            section = None
            for rec in section_rows_sorted:
                if int(rec["row"]) <= int(image_row):
                    section = str(rec.get("section", "Unassigned"))
                else:
                    break
            if section is not None:
                return section
            # If image is above first section, use first section.
            return str(section_rows_sorted[0].get("section", "Unassigned"))

        def _subsection_from_row(section, image_row):
            in_section = [
                r for r in subsection_rows_sorted if str(r.get("section", "")) == str(section)
            ]
            if not in_section:
                subsection_map = target_section_schema.get(section, {})
                return str(next(iter(subsection_map.keys()), "General")) if isinstance(subsection_map, dict) else "General"
            subsection = None
            for rec in in_section:
                if int(rec["row"]) <= int(image_row):
                    subsection = str(rec.get("subsection", "General"))
                else:
                    break
            if subsection is not None:
                return subsection
            return str(in_section[0].get("subsection", "General"))

        def deterministic_assign(image_row):
            if not isinstance(image_row, int) or image_row <= 0:
                return {"section": "Unassigned", "subsection": "General"}
            section = _section_from_row(image_row)
            subsection = _subsection_from_row(section, image_row)
            return {"section": section, "subsection": subsection}

        fallback = {}
        for img in images:
            fallback[str(img.get("image_id"))] = deterministic_assign(img.get("row"))

        llm_images = [
            {
                "image_id": str(img.get("image_id")),
                "row": int(img.get("row", 0)) if isinstance(img.get("row"), int) else 0,
                "col": int(img.get("col", 0)) if isinstance(img.get("col"), int) else 0,
                "anchor_cell": str(img.get("anchor_cell", "")),
            }
            for img in images
        ]

        prompt = (
            "Assign each image to the closest section/subsection using row proximity.\n"
            "Rules:\n"
            "- Prefer closest subsection row when available.\n"
            "- If subsection unclear, pick closest section and a valid subsection under that section.\n"
            "- Use only given section/subsection names.\n"
            "- Keep one assignment per image_id.\n"
            "Return JSON only: {\"assignments\":[{\"image_id\":\"...\",\"section\":\"...\",\"subsection\":\"...\",\"position\":123}]}\n\n"
            + "Sections:\n"
            + json.dumps(section_names, ensure_ascii=False)
            + "\n\nSection rows:\n"
            + json.dumps(section_rows, ensure_ascii=False)
            + "\n\nSubsection rows:\n"
            + json.dumps(subsection_rows, ensure_ascii=False)
            + "\n\nImages:\n"
            + json.dumps(llm_images, ensure_ascii=False)
        )
        parsed = self._call_openai_json(prompt)
        assignments = parsed.get("assignments") if isinstance(parsed, dict) else None
        if not isinstance(assignments, list):
            assignments = []

        out = {}
        for rec in assignments:
            if not isinstance(rec, dict):
                continue
            image_id = str(rec.get("image_id", "")).strip()
            if not image_id:
                continue
            section = str(rec.get("section", "")).strip()
            subsection = str(rec.get("subsection", "")).strip()
            position = rec.get("position")
            if (section, subsection) not in valid_pairs:
                fb = fallback.get(image_id, {"section": "Unassigned", "subsection": "General"})
                section, subsection = fb["section"], fb["subsection"]
            try:
                position_int = int(position)
            except Exception:
                position_int = 0
            out[image_id] = {
                "section": section,
                "subsection": subsection,
                "position": position_int,
            }

        for img in images:
            image_id = str(img.get("image_id"))
            if image_id not in out:
                fb = fallback.get(image_id, {"section": "Unassigned", "subsection": "General"})
                row_num = img.get("row")
                out[image_id] = {
                    "section": fb["section"],
                    "subsection": fb["subsection"],
                    "position": int(row_num) if isinstance(row_num, int) else 0,
                }
        return out

    def downloader_from_datalake_tool(self, tool_input):
        """Download objects from MinIO path into local directory.

        Input: {bucket, prefix, local_dir}
        Return: {downloaded_files, message}
        """
        bucket_name = tool_input["bucket"]
        raw_prefix = tool_input.get("prefix", "")
        local_dir = Path(tool_input.get("local_dir", self.config.local_dir))
        normalized_prefix = raw_prefix.lstrip("/")
        local_dir.mkdir(parents=True, exist_ok=True)

        downloaded_files = []
        for obj in self.client.list_objects(bucket_name, prefix=normalized_prefix, recursive=True):
            object_name = obj.object_name
            if not object_name:
                continue
            relative_path = object_name
            if normalized_prefix and object_name.startswith(normalized_prefix):
                relative_path = object_name[len(normalized_prefix) :].lstrip("/")
            destination = local_dir / relative_path
            destination.parent.mkdir(parents=True, exist_ok=True)
            self.client.fget_object(bucket_name, object_name, str(destination))
            downloaded_files.append(str(destination))

        return {
            "downloaded_files": downloaded_files,
            "message": f"Downloaded {len(downloaded_files)} file(s).",
        }

    def read_excel_rows_text_tool(self, tool_input):
        """Extract Excel rows as plain text with one row per line.

        Input: {excel_files}
        Return: {excel_rows_text, message}
        """
        excel_files = tool_input.get("excel_files", [])
        excel_rows_text = {}
        for file_path in excel_files:
            file_name = Path(file_path).name
            excel_rows_text[file_name] = {"file_path": file_path, "sheets": {}, "error": None}
            try:
                workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    sheet_rows = []
                    for row_values in ws.iter_rows(values_only=True):
                        normalized = []
                        for v in row_values:
                            n = self._normalize_cell_value(v)
                            if n is not None:
                                normalized.append(n)
                        if normalized:
                            sheet_rows.append(normalized)
                    sheet_text = self._rows_to_text(sheet_rows)
                    excel_rows_text[file_name]["sheets"][sheet_name] = sheet_text
            except Exception as exc:
                excel_rows_text[file_name]["error"] = str(exc)
        
        return {
            "excel_rows_text": excel_rows_text,
            "message": f"Loaded row text for {len(excel_rows_text)} file(s).",
        }

    def extract_excel_images_by_section_tool(self, tool_input):
        """Extract embedded Excel images and map them to sections per sheet.

        Input: {excel_files, target_sections, output_dir?}
        Return: {images_by_section, images, output_dir, message}
        """
        excel_files = tool_input.get("excel_files", [])
        target_sections = tool_input.get("target_sections", [])
        output_dir = Path(tool_input.get("output_dir", "output/excel_images"))
        output_dir.mkdir(parents=True, exist_ok=True)

        images_by_section = {}
        images = []
        saved_count = 0

        for file_path in excel_files:
            file_name = Path(file_path).name
            file_stem = Path(file_name).stem
            images_by_section[file_name] = {"file_path": file_path, "sheets": {}, "error": None}
            try:
                workbook = load_workbook(filename=file_path, data_only=True, read_only=False)
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    section_ranges = self._row_section_ranges(ws, target_sections)

                    sheet_map = {}
                    ws_images = list(getattr(ws, "_images", []))
                    for index, img in enumerate(ws_images, start=1):
                        row_num, col_num = self._anchor_to_row_col(getattr(img, "anchor", None))
                        section = self._resolve_section_for_row(row_num, section_ranges)
                        section_key = str(section or "Unassigned")
                        section_folder = (
                            output_dir
                            / self._safe_path_name(file_stem)
                            / self._safe_path_name(sheet_name)
                            / self._safe_path_name(section_key)
                        )
                        section_folder.mkdir(parents=True, exist_ok=True)

                        ext = "png"
                        image_path = getattr(img, "path", None)
                        if image_path:
                            suffix = Path(str(image_path)).suffix.lower().lstrip(".")
                            if suffix:
                                ext = suffix
                        image_data = img._data() if hasattr(img, "_data") else None
                        if not image_data:
                            continue

                        anchor_cell = (
                            f"{get_column_letter(col_num)}{row_num}"
                            if row_num is not None and col_num is not None
                            else "unknown"
                        )
                        file_out = section_folder / f"img_{index}_{anchor_cell}.{ext}"
                        file_out.write_bytes(image_data)

                        rec = {
                            "file": file_name,
                            "sheet": sheet_name,
                            "section": section_key,
                            "anchor_cell": anchor_cell,
                            "row": row_num,
                            "col": col_num,
                            "image_file": str(file_out),
                        }
                        images.append(rec)
                        sheet_map.setdefault(section_key, []).append(rec)
                        saved_count += 1

                    images_by_section[file_name]["sheets"][sheet_name] = sheet_map
            except Exception as exc:
                images_by_section[file_name]["error"] = str(exc)

        return {
            "images_by_section": images_by_section,
            "images": images,
            "output_dir": str(output_dir),
            "message": f"Extracted {saved_count} image(s) from {len(excel_files)} file(s).",
        }

    def arrange_extract_excel_images_llm_tool(self, tool_input):
        """Extract images and assign them to closest section/subsection via LLM + fallback.

        Input: {excel_files, target_section_schema, products_dict?, output_dir?}
        Return: {product_images, message}
        """
        excel_files = tool_input.get("excel_files", [])
        target_section_schema = tool_input.get("target_section_schema", {})
        products_dict = tool_input.get("products_dict", {})
        output_dir = Path(tool_input.get("output_dir", "output/excel_images_llm"))
        output_dir.mkdir(parents=True, exist_ok=True)

        if not isinstance(target_section_schema, dict) or not target_section_schema:
            raise RuntimeError("Image assignment failed: target_section_schema must be a non-empty dict.")

        product_by_file_sheet = {}
        if isinstance(products_dict, dict):
            for product_id, payload in products_dict.items():
                if not isinstance(payload, dict):
                    continue
                product_by_file_sheet[(str(payload.get("file", "")), str(payload.get("sheet", "")))] = str(product_id)

        product_images = []
        processed_images = 0
        for file_path in excel_files:
            file_name = Path(file_path).name
            file_stem = Path(file_name).stem
            try:
                workbook = load_workbook(filename=file_path, data_only=True, read_only=False)
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    section_rows = []
                    subsection_rows = []

                    section_norm = {self._norm_label(k): str(k) for k in target_section_schema.keys()}
                    subsection_norm = {}
                    for section, subsection_map in target_section_schema.items():
                        if not isinstance(subsection_map, dict):
                            continue
                        for subsection in subsection_map.keys():
                            subsection_norm[self._norm_label(subsection)] = (str(section), str(subsection))

                    for row_idx, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
                        tokens = []
                        for value in row_values:
                            normalized = self._normalize_cell_value(value)
                            if normalized:
                                tokens.append(normalized)
                        if not tokens:
                            continue

                        found_section = None
                        for token in tokens:
                            token_norm = self._norm_label(token)
                            if not token_norm:
                                continue
                            if token_norm in section_norm:
                                found_section = section_norm[token_norm]
                                break
                            for k_norm, sec_name in section_norm.items():
                                if len(k_norm) >= 4 and (k_norm in token_norm or token_norm in k_norm):
                                    found_section = sec_name
                                    break
                            if found_section:
                                break
                        if found_section:
                            section_rows.append({"section": found_section, "row": row_idx})

                        found_subsection = None
                        for token in tokens:
                            token_norm = self._norm_label(token)
                            if not token_norm:
                                continue
                            if token_norm in subsection_norm:
                                found_subsection = subsection_norm[token_norm]
                                break
                            for k_norm, sec_sub in subsection_norm.items():
                                if len(k_norm) >= 4 and (k_norm in token_norm or token_norm in k_norm):
                                    found_subsection = sec_sub
                                    break
                            if found_subsection:
                                break
                        if found_subsection:
                            section_name, subsection_name = found_subsection
                            subsection_rows.append(
                                {"section": section_name, "subsection": subsection_name, "row": row_idx}
                            )

                    sheet_images = []
                    ws_images = list(getattr(ws, "_images", []))
                    for index, img in enumerate(ws_images, start=1):
                        row_num, col_num = self._anchor_to_row_col(getattr(img, "anchor", None))
                        ext = "png"
                        image_path = getattr(img, "path", None)
                        if image_path:
                            suffix = Path(str(image_path)).suffix.lower().lstrip(".")
                            if suffix:
                                ext = suffix
                        image_data = img._data() if hasattr(img, "_data") else None
                        if not image_data:
                            continue

                        anchor_cell = (
                            f"{get_column_letter(col_num)}{row_num}"
                            if row_num is not None and col_num is not None
                            else "unknown"
                        )
                        image_id = f"{self._safe_path_name(file_stem)}_{self._safe_path_name(sheet_name)}_{index}"
                        file_out = (
                            output_dir
                            / self._safe_path_name(file_stem)
                            / self._safe_path_name(sheet_name)
                            / f"{image_id}_{anchor_cell}.{ext}"
                        )
                        file_out.parent.mkdir(parents=True, exist_ok=True)
                        file_out.write_bytes(image_data)

                        sheet_images.append(
                            {
                                "image_id": image_id,
                                "row": int(row_num) if isinstance(row_num, int) else 0,
                                "col": int(col_num) if isinstance(col_num, int) else 0,
                                "anchor_cell": anchor_cell,
                                "image_file": str(file_out),
                                "image_blob": base64.b64encode(image_data).decode("ascii"),
                            }
                        )

                    assignments = self._assign_images_with_llm_and_fallback(
                        images=sheet_images,
                        section_rows=section_rows,
                        subsection_rows=subsection_rows,
                        target_section_schema=target_section_schema,
                    )
                    product_id = product_by_file_sheet.get((file_name, sheet_name), Path(file_name).stem.split()[0])

                    for img in sheet_images:
                        image_id = str(img["image_id"])
                        assigned = assignments.get(
                            image_id,
                            {"section": "Unassigned", "subsection": "General", "position": img.get("row", 0)},
                        )
                        product_images.append(
                            {
                                "product_id": str(product_id),
                                "section": str(assigned.get("section", "Unassigned")),
                                "subsection": str(assigned.get("subsection", "General")),
                                "image_id": image_id,
                                "position": int(assigned.get("position", 0)),
                                "image_blob": str(img["image_blob"]),
                                "file": file_name,
                                "sheet": sheet_name,
                                "anchor_cell": str(img["anchor_cell"]),
                                "image_file": str(img["image_file"]),
                            }
                        )
                        processed_images += 1
            except Exception:
                continue

        return {
            "product_images": product_images,
            "message": f"Prepared {processed_images} image assignment(s) for ingestion.",
        }

    def ingest_product_images_clickhouse_tool(self, tool_input):
        """Create product_images table in ClickHouse if needed and ingest image blobs.

        Input: {product_images}
        Return: {inserted_rows, table, run_id, message}
        """
        product_images = tool_input.get("product_images", [])
        db = self.config.clickhouse_db
        table = "product_images"
        run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

        self._clickhouse_http_query(f"CREATE DATABASE IF NOT EXISTS {db}")
        create_sql = f"""
        CREATE TABLE IF NOT EXISTS {db}.{table} (
            run_id String,
            product_id String,
            section String,
            subsection String,
            image_id String,
            position UInt32,
            image_blob String,
            ingested_at DateTime
        ) ENGINE = MergeTree
        ORDER BY (product_id, section, subsection, position, image_id)
        """
        self._clickhouse_http_query(create_sql)

        if not product_images:
            return {
                "inserted_rows": 0,
                "table": f"{db}.{table}",
                "run_id": run_id,
                "message": "No image rows to ingest. Table ensured.",
            }

        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        lines = []
        for row in product_images:
            rec = {
                "run_id": run_id,
                "product_id": str(row.get("product_id", "")),
                "section": str(row.get("section", "Unassigned")),
                "subsection": str(row.get("subsection", "General")),
                "image_id": str(row.get("image_id", "")),
                "position": int(row.get("position", 0)),
                "image_blob": str(row.get("image_blob", "")),
                "ingested_at": now_str,
            }
            lines.append(json.dumps(rec, ensure_ascii=False))

        insert_sql = f"INSERT INTO {db}.{table} FORMAT JSONEachRow"
        self._clickhouse_http_query(insert_sql, body="\n".join(lines))
        return {
            "inserted_rows": len(lines),
            "table": f"{db}.{table}",
            "run_id": run_id,
            "message": f"Ingested {len(lines)} image row(s) into {db}.{table}.",
        }

    def translate_chinese_text_llm_tool(self, tool_input):
        """Translate each sheet text in excel_rows_text and return same dict shape.

        Input: {excel_rows_text}
        Return: {excel_rows_text, message}
        """
        excel_rows_text = tool_input.get("excel_rows_text", {})
        translated_all = {}
        for file_name, payload in excel_rows_text.items():
            translated_payload = {
                "file_path": payload.get("file_path"),
                "sheets": {},
                "error": payload.get("error"),
            }
            for sheet_name, sheet_text in payload.get("sheets", {}).items():
                translated_payload["sheets"][sheet_name] = self._translate_chinese_to_english(
                    str(sheet_text)
                )
                
            translated_all[file_name] = translated_payload
        return {
            "excel_rows_text": translated_all,
            "message": f"Translated sheet text for {len(translated_all)} file(s).",
        }

    def fill_schema_values_llm_tool(self, tool_input):
        """LLM Tool #3: fill schema key-values per product/sheet.

        Input: {excel_rows_text, target_sections, target_section_schema}
        Return: {products_dict, product_table_rows, message}
        Strict LLM-only extraction; raises on invalid/empty LLM output.
        """
        excel_rows_text = tool_input.get("excel_rows_text", {})
        target_sections = tool_input.get("target_sections", [])
        target_section_schema = tool_input.get("target_section_schema", {})
        if not isinstance(target_sections, list) or not target_sections:
            raise RuntimeError("LLM extraction failed: target_sections must be a non-empty list.")
        if not isinstance(target_section_schema, dict) or not target_section_schema:
            raise RuntimeError("LLM extraction failed: target_section_schema must be a non-empty dict.")
        if self._get_llm() is None:
            raise RuntimeError("LLM extraction failed: OpenAI is not configured or unavailable.")
        products = {}
        table_rows = []

        for file_name, payload in excel_rows_text.items():
            default_id = Path(file_name).stem.split()[0]
            for sheet_name, sheet_text in payload.get("sheets", {}).items():
                rows = self._text_to_rows(sheet_text)
                product_id = self._extract_product_id(rows, default_id)
                prompt = (
                    self.prompts["extract_key_value_pairs_by_sections"]
                    + "\n\nTarget sections:\n"
                    + json.dumps(target_sections, ensure_ascii=False)
                    + "\n\nExpected section/subsection/key schema (return these keys; use 'not available' when missing):\n"
                    + json.dumps(target_section_schema, ensure_ascii=False)
                    + "\n\nRows:\n"
                    + sheet_text
                )
                parsed = self._call_openai_json(prompt)
                sections_filled, parsed_product_id = self._normalize_sections_from_llm(
                    parsed, target_section_schema
                )
                if sections_filled is None:
                    repair_prompt = (
                        prompt
                        + "\n\nYour previous output was not valid. Return ONLY valid JSON with top-level keys "
                        + "\"product_id\" and \"sections\" where \"sections\" is an object."
                    )
                    parsed = self._call_openai_json(repair_prompt)
                    sections_filled, parsed_product_id = self._normalize_sections_from_llm(
                        parsed, target_section_schema
                    )
                if sections_filled is None:
                    raise RuntimeError(
                        f"LLM extraction failed for file '{file_name}' sheet '{sheet_name}': invalid JSON sections."
                    )
                pid = parsed_product_id or str(product_id)

                normalized_sections = {}
                for section in target_sections:
                    expected_subsections = target_section_schema.get(section, {})
                    current_section = sections_filled.get(section, {})
                    if not isinstance(current_section, dict):
                        current_section = {}
                    section_out = {}
                    for subsection, expected_keys in expected_subsections.items():
                        current_sub = current_section.get(subsection, {})
                        if not isinstance(current_sub, dict):
                            # Case-insensitive subsection lookup.
                            matched = None
                            for sub_name, sub_payload in current_section.items():
                                if str(sub_name).strip().lower() == str(subsection).strip().lower():
                                    matched = sub_payload
                                    break
                            current_sub = matched if isinstance(matched, dict) else {}
                        out = {}
                        current_lc = {str(k).strip().lower(): v for k, v in current_sub.items()}
                        for expected_key in expected_keys:
                            value = current_sub.get(expected_key)
                            if value is None:
                                value = current_lc.get(str(expected_key).strip().lower())
                            if value is None or str(value).strip() == "":
                                value = "not available"
                            out[str(expected_key)] = str(value)
                        section_out[str(subsection)] = out
                    normalized_sections[section] = section_out
                sections_filled = normalized_sections

                pair_count = 0
                for subsection_map in sections_filled.values():
                    if not isinstance(subsection_map, dict):
                        continue
                    for kv in subsection_map.values():
                        if isinstance(kv, dict):
                            pair_count += len(kv)
                if pair_count == 0:
                    raise RuntimeError(
                        f"LLM extraction failed for file '{file_name}' sheet '{sheet_name}': extracted 0 key-value pairs."
                    )

                sections_filled = self._translate_sections_to_english(sections_filled)

                products[pid] = {
                    "product_id": pid,
                    "file": file_name,
                    "sheet": sheet_name,
                    "sections": sections_filled,
                }
                for sec, subsection_map in sections_filled.items():
                    if not isinstance(subsection_map, dict):
                        continue
                    for subsection, kv in subsection_map.items():
                        if not isinstance(kv, dict):
                            continue
                        for key, value in kv.items():
                            table_rows.append(
                                {
                                    "product_id": pid,
                                    "file": file_name,
                                    "sheet": sheet_name,
                                    "section": sec,
                                    "subsection": subsection,
                                    "key": key,
                                    "value": value,
                                }
                            )

        return {
            "products_dict": products,
            "product_table_rows": table_rows,
            "message": f"Filled schema values for {len(products)} product(s).",
        }

    def ingest_product_facts_clickhouse_tool(self, tool_input):
        """Create product_facts table in ClickHouse if needed and ingest rows.

        Input: {product_table_rows}
        Return: {inserted_rows, table, run_id, message}
        """
        product_table_rows = tool_input.get("product_table_rows", [])
        db = self.config.clickhouse_db
        table = "product_facts"
        run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

        self._clickhouse_http_query(f"CREATE DATABASE IF NOT EXISTS {db}")
        create_sql = f"""
        CREATE TABLE IF NOT EXISTS {db}.{table} (
            run_id String,
            product_id String,
            file String,
            sheet String,
            section String,
            subsection String,
            fact_key String,
            fact_value String,
            ingested_at DateTime
        ) ENGINE = MergeTree
        ORDER BY (product_id, section, subsection, fact_key, sheet, file)
        """
        self._clickhouse_http_query(create_sql)

        if not product_table_rows:
            return {
                "inserted_rows": 0,
                "table": f"{db}.{table}",
                "run_id": run_id,
                "message": "No rows to ingest. Table ensured.",
            }

        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        lines = []
        for row in product_table_rows:
            rec = {
                "run_id": run_id,
                "product_id": str(row.get("product_id", "")),
                "file": str(row.get("file", "")),
                "sheet": str(row.get("sheet", "")),
                "section": str(row.get("section", "")),
                "subsection": str(row.get("subsection", "General")),
                "fact_key": str(row.get("key", "")),
                "fact_value": str(row.get("value", "")),
                "ingested_at": now_str,
            }
            lines.append(json.dumps(rec, ensure_ascii=False))

        insert_sql = f"INSERT INTO {db}.{table} FORMAT JSONEachRow"
        self._clickhouse_http_query(insert_sql, body="\n".join(lines))
        return {
            "inserted_rows": len(lines),
            "table": f"{db}.{table}",
            "run_id": run_id,
            "message": f"Ingested {len(lines)} row(s) into {db}.{table}.",
        }

    def clickhouse_table_count_tool(self, tool_input):
        """Return row count for a ClickHouse table in configured DB.

        Input: {table}
        Return: {table, row_count, message}
        """
        table = str(tool_input.get("table", "product_facts"))
        db = self.config.clickhouse_db
        query = f"SELECT count() AS c FROM {db}.{table} FORMAT JSON"
        raw = self._clickhouse_http_query(query)
        try:
            parsed = json.loads(raw)
            rows = parsed.get("data", [])
            count = int(rows[0].get("c", 0)) if rows else 0
        except Exception as exc:
            raise RuntimeError(f"Failed to parse ClickHouse count response: {exc}") from exc
        return {
            "table": f"{db}.{table}",
            "row_count": count,
            "message": f"Table {db}.{table} has {count} row(s).",
        }

    def registry(self):
        """Return tool registry consumed by the agent graph."""
        return {
            "downloader_from_datalake": self.downloader_from_datalake_tool,
            "read_excel_rows_text": self.read_excel_rows_text_tool,
            "extract_excel_images_by_section": self.extract_excel_images_by_section_tool,
            "arrange_extract_excel_images_llm": self.arrange_extract_excel_images_llm_tool,
            "translate_chinese_text_llm": self.translate_chinese_text_llm_tool,
            "fill_schema_values_llm": self.fill_schema_values_llm_tool,
            "ingest_product_facts_clickhouse": self.ingest_product_facts_clickhouse_tool,
            "ingest_product_images_clickhouse": self.ingest_product_images_clickhouse_tool,
            "clickhouse_table_count": self.clickhouse_table_count_tool,
        }
