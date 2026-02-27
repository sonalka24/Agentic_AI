import base64
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from extract_data_agent.telemetry import logger


class LlmToolsMixin:
    def translate_chinese_text_llm_tool(self, tool_input):
        excel_rows_text = tool_input.get("excel_rows_text", {})
        translated_all = {}
        for file_name, payload in excel_rows_text.items():
            translated_payload = {"file_path": payload.get("file_path"), "sheets": {}, "error": payload.get("error")}
            for sheet_name, sheet_text in payload.get("sheets", {}).items():
                contains_chinese = self._contains_chinese(str(sheet_text))
                logger.info("Translation step: file=%s sheet=%s contains_chinese=%s", file_name, sheet_name, contains_chinese)
                translated_payload["sheets"][sheet_name] = self._translate_chinese_to_english(str(sheet_text))
            translated_all[file_name] = translated_payload
        return {"excel_rows_text": translated_all, "message": f"Translated sheet text for {len(translated_all)} file(s)."}

    def fill_schema_values_llm_tool(self, tool_input):
        excel_rows_text = tool_input.get("excel_rows_text", {})
        target_sections = tool_input.get("target_sections", [])
        target_section_schema = tool_input.get("target_section_schema", {})
        if not isinstance(target_sections, list) or not target_sections:
            raise RuntimeError("LLM extraction failed: target_sections must be a non-empty list.")
        if not isinstance(target_section_schema, dict) or not target_section_schema:
            raise RuntimeError("LLM extraction failed: target_section_schema must be a non-empty dict.")
        if self._get_llm() is None:
            raise RuntimeError("LLM extraction failed: OpenAI is not configured or unavailable.")
        products = {}
        table_rows = []
        for file_name, payload in excel_rows_text.items():
            default_id = Path(file_name).stem.split()[0]
            for sheet_name, sheet_text in payload.get("sheets", {}).items():
                logger.info("Schema fill step: file=%s sheet=%s phase=prepare", file_name, sheet_name)
                rows = self._text_to_rows(sheet_text)
                product_id = self._extract_product_id(rows, default_id)
                prompt = self.prompts["extract_key_value_pairs_by_sections"] + "\n\nTarget sections:\n" + json.dumps(target_sections, ensure_ascii=False) + "\n\nExpected section/subsection/key schema (return these keys; use 'not available' when missing):\n" + json.dumps(target_section_schema, ensure_ascii=False) + "\n\nRows:\n" + sheet_text
                logger.info("Schema fill step: file=%s sheet=%s phase=llm_extract", file_name, sheet_name)
                parsed = self._call_openai_json(prompt)
                sections_filled, parsed_product_id = self._normalize_sections_from_llm(parsed, target_section_schema)
                if sections_filled is None:
                    repair_prompt = (
                        prompt
                        + "\n\nYour previous output was not valid. Return ONLY valid JSON with top-level keys "
                        + "\"product_id\" and \"sections\" where \"sections\" is an object."
                    )
                    logger.info("Schema fill step: file=%s sheet=%s phase=llm_repair", file_name, sheet_name)
                    parsed = self._call_openai_json(repair_prompt)
                    sections_filled, parsed_product_id = self._normalize_sections_from_llm(parsed, target_section_schema)
                if sections_filled is None:
                    raise RuntimeError(f"LLM extraction failed for file '{file_name}' sheet '{sheet_name}': invalid JSON sections.")
                pid = parsed_product_id or str(product_id)
                normalized_sections = {}
                for section in target_sections:
                    expected_subsections = target_section_schema.get(section, {})
                    current_section = sections_filled.get(section, {})
                    if not isinstance(current_section, dict):
                        current_section = {}
                    section_out = {}
                    for subsection, expected_keys in expected_subsections.items():
                        current_sub = current_section.get(subsection, {})
                        if not isinstance(current_sub, dict):
                            matched = None
                            for sub_name, sub_payload in current_section.items():
                                if str(sub_name).strip().lower() == str(subsection).strip().lower():
                                    matched = sub_payload
                                    break
                            current_sub = matched if isinstance(matched, dict) else {}
                        out = {}
                        current_lc = {str(k).strip().lower(): v for k, v in current_sub.items()}
                        for expected_key in expected_keys:
                            value = current_sub.get(expected_key)
                            if value is None:
                                value = current_lc.get(str(expected_key).strip().lower())
                            if value is None or str(value).strip() == "":
                                value = "not available"
                            out[str(expected_key)] = str(value)
                        section_out[str(subsection)] = out
                    normalized_sections[section] = section_out
                sections_filled = normalized_sections
                pair_count = 0
                for subsection_map in sections_filled.values():
                    if not isinstance(subsection_map, dict):
                        continue
                    for kv in subsection_map.values():
                        if isinstance(kv, dict):
                            pair_count += len(kv)
                if pair_count == 0:
                    raise RuntimeError(f"LLM extraction failed for file '{file_name}' sheet '{sheet_name}': extracted 0 key-value pairs.")
                logger.info("Schema fill step: file=%s sheet=%s phase=translate_sections", file_name, sheet_name)
                sections_filled = self._translate_sections_to_english(sections_filled)
                logger.info("Schema fill step: file=%s sheet=%s phase=complete product_id=%s rows=%s", file_name, sheet_name, pid, len(rows))
                products[pid] = {"product_id": pid, "file": file_name, "sheet": sheet_name, "sections": sections_filled}
                for sec, subsection_map in sections_filled.items():
                    if not isinstance(subsection_map, dict):
                        continue
                    for subsection, kv in subsection_map.items():
                        if not isinstance(kv, dict):
                            continue
                        for key, value in kv.items():
                            table_rows.append({"product_id": pid, "file": file_name, "sheet": sheet_name, "section": sec, "subsection": subsection, "key": key, "value": value})
        return {"products_dict": products, "product_table_rows": table_rows, "message": f"Filled schema values for {len(products)} product(s)."}

    def arrange_extract_excel_images_llm_tool(self, tool_input):
        excel_files = tool_input.get("excel_files", [])
        target_section_schema = tool_input.get("target_section_schema", {})
        products_dict = tool_input.get("products_dict", {})
        output_dir = Path(tool_input.get("output_dir", "output/excel_images_llm"))
        output_dir.mkdir(parents=True, exist_ok=True)
        if not isinstance(target_section_schema, dict) or not target_section_schema:
            raise RuntimeError("Image assignment failed: target_section_schema must be a non-empty dict.")
        product_by_file_sheet = {}
        if isinstance(products_dict, dict):
            for product_id, payload in products_dict.items():
                if isinstance(payload, dict):
                    product_by_file_sheet[(str(payload.get("file", "")), str(payload.get("sheet", "")))] = str(product_id)
        product_images = []
        processed_images = 0
        for file_path in excel_files:
            file_name = Path(file_path).name
            file_stem = Path(file_name).stem
            logger.info("Arrange images step: file=%s phase=start", file_name)
            try:
                workbook = load_workbook(filename=file_path, data_only=True, read_only=False)
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    logger.info("Arrange images step: file=%s sheet=%s phase=scan_sheet", file_name, sheet_name)
                    section_rows = []
                    subsection_rows = []
                    section_norm = {self._norm_label(k): str(k) for k in target_section_schema.keys()}
                    subsection_norm = {}
                    for section, subsection_map in target_section_schema.items():
                        if isinstance(subsection_map, dict):
                            for subsection in subsection_map.keys():
                                subsection_norm[self._norm_label(subsection)] = (str(section), str(subsection))
                    for row_idx, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
                        tokens = []
                        for value in row_values:
                            normalized = self._normalize_cell_value(value)
                            if normalized:
                                tokens.append(normalized)
                        if not tokens:
                            continue
                        found_section = None
                        for token in tokens:
                            token_norm = self._norm_label(token)
                            if not token_norm:
                                continue
                            if token_norm in section_norm:
                                found_section = section_norm[token_norm]
                                break
                            for k_norm, sec_name in section_norm.items():
                                if len(k_norm) >= 4 and (k_norm in token_norm or token_norm in k_norm):
                                    found_section = sec_name
                                    break
                            if found_section:
                                break
                        if found_section:
                            section_rows.append({"section": found_section, "row": row_idx})
                        found_subsection = None
                        for token in tokens:
                            token_norm = self._norm_label(token)
                            if not token_norm:
                                continue
                            if token_norm in subsection_norm:
                                found_subsection = subsection_norm[token_norm]
                                break
                            for k_norm, sec_sub in subsection_norm.items():
                                if len(k_norm) >= 4 and (k_norm in token_norm or token_norm in k_norm):
                                    found_subsection = sec_sub
                                    break
                            if found_subsection:
                                break
                        if found_subsection:
                            section_name, subsection_name = found_subsection
                            subsection_rows.append({"section": section_name, "subsection": subsection_name, "row": row_idx})
                    logger.info(
                        "Arrange images step: file=%s sheet=%s phase=detect_rows section_rows=%s subsection_rows=%s",
                        file_name,
                        sheet_name,
                        len(section_rows),
                        len(subsection_rows),
                    )
                    sheet_images = []
                    for index, img in enumerate(list(getattr(ws, "_images", [])), start=1):
                        row_num, col_num = self._anchor_to_row_col(getattr(img, "anchor", None))
                        ext = "png"
                        image_path = getattr(img, "path", None)
                        if image_path:
                            suffix = Path(str(image_path)).suffix.lower().lstrip(".")
                            if suffix:
                                ext = suffix
                        image_data = img._data() if hasattr(img, "_data") else None
                        if not image_data:
                            continue
                        anchor_cell = f"{get_column_letter(col_num)}{row_num}" if row_num is not None and col_num is not None else "unknown"
                        image_id = f"{self._safe_path_name(file_stem)}_{self._safe_path_name(sheet_name)}_{index}"
                        file_out = output_dir / self._safe_path_name(file_stem) / self._safe_path_name(sheet_name) / f"{image_id}_{anchor_cell}.{ext}"
                        file_out.parent.mkdir(parents=True, exist_ok=True)
                        file_out.write_bytes(image_data)
                        sheet_images.append({"image_id": image_id, "row": int(row_num) if isinstance(row_num, int) else 0, "col": int(col_num) if isinstance(col_num, int) else 0, "anchor_cell": anchor_cell, "image_file": str(file_out), "image_blob": base64.b64encode(image_data).decode("ascii")})
                    logger.info(
                        "Arrange images step: file=%s sheet=%s phase=extract_images image_count=%s",
                        file_name,
                        sheet_name,
                        len(sheet_images),
                    )
                    logger.info("Arrange images step: file=%s sheet=%s phase=assign_llm", file_name, sheet_name)
                    assignments = self._assign_images_with_llm_and_fallback(sheet_images, section_rows, subsection_rows, target_section_schema)
                    product_id = product_by_file_sheet.get((file_name, sheet_name), Path(file_name).stem.split()[0])
                    logger.info(
                        "Arrange images step: file=%s sheet=%s phase=assignment_complete assigned=%s",
                        file_name,
                        sheet_name,
                        len(assignments),
                    )
                    for img in sheet_images:
                        image_id = str(img["image_id"])
                        assigned = assignments.get(image_id, {"section": "Unassigned", "subsection": "General", "position": img.get("row", 0)})
                        product_images.append({"product_id": str(product_id), "section": str(assigned.get("section", "Unassigned")), "subsection": str(assigned.get("subsection", "General")), "image_id": image_id, "position": int(assigned.get("position", 0)), "image_blob": str(img["image_blob"]), "file": file_name, "sheet": sheet_name, "anchor_cell": str(img["anchor_cell"]), "image_file": str(img["image_file"])})
                        processed_images += 1
                    logger.info(
                        "Arrange images step: file=%s sheet=%s phase=complete product_id=%s image_count=%s",
                        file_name,
                        sheet_name,
                        product_id,
                        len(sheet_images),
                    )
                logger.info("Arrange images step: file=%s phase=file_complete processed_images=%s", file_name, processed_images)
            except Exception as exc:
                logger.exception("Arrange images step failed: file=%s error=%s", file_name, exc)
                continue
        return {"product_images": product_images, "message": f"Prepared {processed_images} image assignment(s) for ingestion."}
