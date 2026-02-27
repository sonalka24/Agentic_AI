from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import config


class FileToolsMixin:
    def downloader_from_datalake_tool(self, tool_input):
        bucket_name = tool_input["bucket"]
        raw_prefix = tool_input.get("prefix", "")
        local_dir = Path(tool_input.get("local_dir", config.local_dir))
        normalized_prefix = raw_prefix.lstrip("/")
        local_dir.mkdir(parents=True, exist_ok=True)
        downloaded_files = []
        for obj in self.client.list_objects(bucket_name, prefix=normalized_prefix, recursive=True):
            object_name = obj.object_name
            if not object_name:
                continue
            relative_path = object_name
            if normalized_prefix and object_name.startswith(normalized_prefix):
                relative_path = object_name[len(normalized_prefix) :].lstrip("/")
            destination = local_dir / relative_path
            destination.parent.mkdir(parents=True, exist_ok=True)
            self.client.fget_object(bucket_name, object_name, str(destination))
            downloaded_files.append(str(destination))
        return {"downloaded_files": downloaded_files, "message": f"Downloaded {len(downloaded_files)} file(s)."}

    def read_excel_rows_text_tool(self, tool_input):
        excel_files = tool_input.get("excel_files", [])
        excel_rows_text = {}
        for file_path in excel_files:
            file_name = Path(file_path).name
            excel_rows_text[file_name] = {"file_path": file_path, "sheets": {}, "error": None}
            try:
                workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    sheet_rows = []
                    for row_values in ws.iter_rows(values_only=True):
                        normalized = []
                        for v in row_values:
                            n = self._normalize_cell_value(v)
                            if n is not None:
                                normalized.append(n)
                        if normalized:
                            sheet_rows.append(normalized)
                    excel_rows_text[file_name]["sheets"][sheet_name] = self._rows_to_text(sheet_rows)
            except Exception as exc:
                excel_rows_text[file_name]["error"] = str(exc)
        return {"excel_rows_text": excel_rows_text, "message": f"Loaded row text for {len(excel_rows_text)} file(s)."}

    def extract_excel_images_by_section_tool(self, tool_input):
        excel_files = tool_input.get("excel_files", [])
        target_sections = tool_input.get("target_sections", [])
        output_dir = Path(tool_input.get("output_dir", "output/excel_images"))
        output_dir.mkdir(parents=True, exist_ok=True)
        images_by_section = {}
        images = []
        saved_count = 0
        for file_path in excel_files:
            file_name = Path(file_path).name
            file_stem = Path(file_name).stem
            images_by_section[file_name] = {"file_path": file_path, "sheets": {}, "error": None}
            try:
                workbook = load_workbook(filename=file_path, data_only=True, read_only=False)
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    section_ranges = self._row_section_ranges(ws, target_sections)
                    sheet_map = {}
                    for index, img in enumerate(list(getattr(ws, "_images", [])), start=1):
                        row_num, col_num = self._anchor_to_row_col(getattr(img, "anchor", None))
                        section_key = str(self._resolve_section_for_row(row_num, section_ranges) or "Unassigned")
                        section_folder = output_dir / self._safe_path_name(file_stem) / self._safe_path_name(sheet_name) / self._safe_path_name(section_key)
                        section_folder.mkdir(parents=True, exist_ok=True)
                        ext = "png"
                        image_path = getattr(img, "path", None)
                        if image_path:
                            suffix = Path(str(image_path)).suffix.lower().lstrip(".")
                            if suffix:
                                ext = suffix
                        image_data = img._data() if hasattr(img, "_data") else None
                        if not image_data:
                            continue
                        anchor_cell = f"{get_column_letter(col_num)}{row_num}" if row_num is not None and col_num is not None else "unknown"
                        file_out = section_folder / f"img_{index}_{anchor_cell}.{ext}"
                        file_out.write_bytes(image_data)
                        rec = {"file": file_name, "sheet": sheet_name, "section": section_key, "anchor_cell": anchor_cell, "row": row_num, "col": col_num, "image_file": str(file_out)}
                        images.append(rec)
                        sheet_map.setdefault(section_key, []).append(rec)
                        saved_count += 1
                    images_by_section[file_name]["sheets"][sheet_name] = sheet_map
            except Exception as exc:
                images_by_section[file_name]["error"] = str(exc)
        return {"images_by_section": images_by_section, "images": images, "output_dir": str(output_dir), "message": f"Extracted {saved_count} image(s) from {len(excel_files)} file(s)."}
